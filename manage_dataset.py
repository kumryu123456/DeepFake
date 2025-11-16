"""
Dataset Management CLI Tool
Manages video-to-audio conversion and segmentation workflow with Excel tracking
"""

import os
import sys
import argparse
import subprocess
from pathlib import Path
from datetime import datetime
import json
import re
import numpy as np
import soundfile as sf
import shutil

# Excel handling
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("ERROR: openpyxl not installed. Please install: pip install openpyxl")
    sys.exit(1)

# TEN VAD
try:
    from ten_vad import TenVad
except ImportError:
    print("WARNING: ten_vad not installed. Mode 2 will not work.")
    print("Install: pip install git+https://github.com/TEN-framework/ten-vad.git")

# pyannote (optional for advanced verification)
try:
    import torch
    from pyannote.audio import Pipeline
    PYANNOTE_AVAILABLE = True
except ImportError:
    PYANNOTE_AVAILABLE = False


def normalize_model(model_str):
    """
    Normalize model name for consistent directory naming
    Examples:
        "Sora 2" -> "Sora"
        "sora 2" -> "Sora"
        "Veo 3" -> "Veo3"
        "veo3" -> "Veo3"
    """
    if not model_str:
        return ""

    # Convert to string and strip
    model = str(model_str).strip()

    # Common normalizations
    model_lower = model.lower()

    if "sora" in model_lower:
        return "Sora"
    elif "veo" in model_lower:
        return "Veo3"

    # Fallback: remove spaces and numbers at the end
    model = re.sub(r'\s+\d+$', '', model)  # Remove trailing space+numbers
    model = model.strip()

    # Capitalize first letter
    if model:
        model = model[0].upper() + model[1:]

    return model


class DatasetManager:
    """Manages the complete dataset workflow"""

    def __init__(self):
        self.input_videos_dir = Path("data/input_videos")
        self.converted_wav_dir = Path("data/converted_wav")
        self.output_segments_dir = Path("data/output_segments")
        self.input_metadata_path = Path("data/input_metadata.xlsx")
        self.output_metadata_path = Path("data/output_metadata.xlsx")

        # Create directories
        self.input_videos_dir.mkdir(parents=True, exist_ok=True)
        self.converted_wav_dir.mkdir(parents=True, exist_ok=True)
        self.output_segments_dir.mkdir(parents=True, exist_ok=True)

    def mode1_prepare(self):
        """Mode 1: Scan mp4 files + Create input_metadata.xlsx + Convert to WAV"""
        print("\n" + "="*60)
        print("MODE 1: PREPARE - MP4 to WAV Conversion")
        print("="*60 + "\n")

        # Scan for mp4 files
        mp4_files = sorted(self.input_videos_dir.glob("*.mp4"))

        if not mp4_files:
            print(f"ERROR: No MP4 files found in {self.input_videos_dir}")
            print("Please add MP4 files to data/input_videos/ and try again.")
            return False

        print(f"Found {len(mp4_files)} MP4 files\n")

        # Create or load input_metadata.xlsx
        if self.input_metadata_path.exists():
            print(f"Loading existing metadata: {self.input_metadata_path}")
            wb = load_workbook(self.input_metadata_path)
            ws = wb.active
            existing_files = {ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)}
        else:
            print(f"Creating new metadata: {self.input_metadata_path}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Input Metadata"

            # Create headers
            headers = ["file_name", "file_last_4", "url", "model", "initial_sample_num"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

            existing_files = set()

        # Process each mp4 file
        new_files_added = 0
        converted_count = 0

        for mp4_path in mp4_files:
            filename = mp4_path.name
            file_last_4 = filename[-8:-4]  # Get last 4 chars before .mp4

            # Check if already in metadata
            if filename in existing_files:
                print(f"[SKIP] {filename} - Already in metadata")
                continue

            # Add to metadata
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1).value = filename
            ws.cell(row=new_row, column=2).value = file_last_4
            ws.cell(row=new_row, column=3).value = ""  # URL - user will fill
            ws.cell(row=new_row, column=4).value = ""  # model - user will fill
            ws.cell(row=new_row, column=5).value = ""  # initial_sample_num - user will fill

            new_files_added += 1
            print(f"[NEW] {filename} (last_4: {file_last_4})")

            # Convert to WAV
            wav_filename = mp4_path.stem + ".wav"
            wav_path = self.converted_wav_dir / wav_filename

            if wav_path.exists():
                print(f"  → WAV already exists: {wav_filename}")
            else:
                success = self._convert_mp4_to_wav(mp4_path, wav_path)
                if success:
                    converted_count += 1
                    print(f"  → [OK] Converted to WAV: {wav_filename}")
                else:
                    print(f"  → [FAIL] Conversion failed")

        # Save metadata
        wb.save(self.input_metadata_path)

        # Summary
        print("\n" + "="*60)
        print("MODE 1 COMPLETE")
        print("="*60)
        print(f"New files added to metadata: {new_files_added}")
        print(f"WAV files converted: {converted_count}")
        print(f"Metadata saved: {self.input_metadata_path}")
        print("\n" + "="*60)
        print("NEXT STEPS:")
        print("1. Open the Excel file and fill in:")
        print("   - url: Source URL of the video (optional)")
        print("   - model: Veo3 or Sora")
        print("   - initial_sample_num: Starting number (only for first file of each model)")
        print("2. Run 'python manage_dataset.py segment' to process segments")
        print("="*60 + "\n")

        return True

    def _convert_mp4_to_wav(self, mp4_path, wav_path, sample_rate=16000):
        """Convert MP4 to WAV using ffmpeg"""
        try:
            # Use ffmpeg to convert
            command = [
                "ffmpeg",
                "-i", str(mp4_path),
                "-vn",  # No video
                "-acodec", "pcm_s16le",  # PCM 16-bit
                "-ar", str(sample_rate),  # Sample rate
                "-ac", "1",  # Mono
                "-y",  # Overwrite
                str(wav_path)
            ]

            result = subprocess.run(
                command,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )

            return result.returncode == 0

        except FileNotFoundError:
            print("  ERROR: ffmpeg not found. Please install ffmpeg:")
            print("  Windows: https://www.ffmpeg.org/download.html")
            print("  Or use: pip install ffmpeg-python")
            return False
        except Exception as e:
            print(f"  ERROR: {e}")
            return False

    def mode2_segment(self, min_duration=4.0, max_duration=10.0, no_cleanup=False):
        """Mode 2: Segment WAV files + Create output_metadata.xlsx + Cleanup"""
        print("\n" + "="*60)
        print("MODE 2: SEGMENT - WAV Segmentation")
        print("="*60 + "\n")

        # Check input_metadata.xlsx exists
        if not self.input_metadata_path.exists():
            print(f"ERROR: {self.input_metadata_path} not found")
            print("Please run 'python manage_dataset.py prepare' first")
            return False

        # Load input metadata
        wb = load_workbook(self.input_metadata_path)
        ws = wb.active

        # Parse metadata WITHOUT assigning sample_num yet
        metadata_records = []
        model_initial_nums = {}  # Track initial_sample_num per model

        for row in range(2, ws.max_row + 1):
            file_name = ws.cell(row=row, column=1).value
            file_last_4 = ws.cell(row=row, column=2).value
            url = ws.cell(row=row, column=3).value or ""
            model_raw = ws.cell(row=row, column=4).value
            initial_sample_num = ws.cell(row=row, column=5).value

            # Validation
            if not file_name:
                continue

            if not model_raw:
                print(f"[SKIP] {file_name} - Missing model")
                continue

            # Normalize model
            model = normalize_model(model_raw)

            # Store initial_sample_num for the model
            if initial_sample_num:
                model_initial_nums[model] = int(initial_sample_num) - 1  # -1 because we'll increment before use

            # Check if model has initial number
            if model not in model_initial_nums:
                print(f"[SKIP] {file_name} - No initial_sample_num for model '{model}'")
                continue

            metadata_records.append({
                "file_name": file_name,
                "file_last_4": file_last_4,
                "url": url,
                "model": model
            })

        if not metadata_records:
            print("ERROR: No valid metadata records found")
            print("Please fill in 'model' and 'initial_sample_num' columns in the Excel file")
            return False

        print(f"Found {len(metadata_records)} valid records\n")

        # Initialize TEN VAD segmenter
        print("Loading TEN VAD model...")
        try:
            vad = TenVad()
            print("TEN VAD loaded successfully\n")
        except Exception as e:
            print(f"ERROR: Failed to load TEN VAD: {e}")
            return False

        # Create output metadata workbook
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.title = "Segment Metadata"

        # Headers - NEW ORDER: file_name, file_last_4, sample_num, segment_num, path, url
        headers = ["file_name", "file_last_4", "sample_num", "segment_num", "path", "url"]
        for col, header in enumerate(headers, start=1):
            cell = output_ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # Process each file - assign sample_num ONLY when segments are created
        total_segments = 0
        output_row = 2
        model_counters = {}  # Track current sample_num per model

        for record in metadata_records:
            file_name = record["file_name"]
            model = record["model"]
            wav_filename = Path(file_name).stem + ".wav"
            wav_path = self.converted_wav_dir / wav_filename

            if not wav_path.exists():
                print(f"[ERROR] {wav_filename} not found in {self.converted_wav_dir}")
                continue

            # Initialize counter for this model if not exists
            if model not in model_counters:
                model_counters[model] = model_initial_nums.get(model, 0)

            print(f"Processing: {wav_filename} (Model: {model})")

            # Segment audio
            try:
                # First, try to generate segments (WITHOUT saving yet)
                segments_data = self._segment_audio_with_vad_preview(
                    wav_path,
                    vad,
                    min_duration,
                    max_duration
                )

                if not segments_data:
                    print(f"  → Generated 0 segments - SKIPPED (no sample_num assigned)")
                    continue

                # Segments exist! Increment counter and assign sample_num
                model_counters[model] += 1
                sample_num = model_counters[model]

                print(f"  → Generated {len(segments_data)} segments - Assigned sample_num: {sample_num}")

                # Now save the segments with the assigned sample_num
                segments = self._save_segments(
                    segments_data,
                    model,
                    sample_num
                )

                # Add to output metadata
                for seg_idx, seg_path in enumerate(segments, start=1):
                    output_ws.cell(row=output_row, column=1).value = file_name
                    output_ws.cell(row=output_row, column=2).value = record["file_last_4"]
                    output_ws.cell(row=output_row, column=3).value = sample_num
                    output_ws.cell(row=output_row, column=4).value = seg_idx
                    output_ws.cell(row=output_row, column=5).value = ""  # path - empty (user will fill with formula)
                    output_ws.cell(row=output_row, column=6).value = record["url"] if seg_idx == 1 else ""
                    output_row += 1

                total_segments += len(segments)

            except Exception as e:
                print(f"  → [ERROR] {e}")
                continue

        # Save output metadata
        output_wb.save(self.output_metadata_path)

        print("\n" + "="*60)
        print("SEGMENTATION COMPLETE")
        print("="*60)
        print(f"Total segments generated: {total_segments}")
        print(f"Output metadata saved: {self.output_metadata_path}")
        print("\n" + "="*60)
        print("GOOGLE SHEETS FORMULA:")
        print('In Path column (E2), use formula:')
        print('  =C2&"/"&D2&"/"&E2&".wav"')
        print('  Example: Sora/1/1.wav')
        print("="*60 + "\n")

        # Cleanup
        if not no_cleanup:
            self._cleanup_workflow()
        else:
            print("Cleanup skipped (--no-cleanup flag)")

        return True

    def _segment_audio_with_vad_preview(self, audio_path, vad, min_duration, max_duration):
        """Preview segments without saving - returns segment data"""
        # Load audio
        audio, sr = sf.read(audio_path)

        # Convert to mono if stereo
        if len(audio.shape) > 1:
            audio = np.mean(audio, axis=1)

        # Convert to int16 for TEN VAD
        audio_int16 = (audio * 32767).astype(np.int16)

        # Run VAD
        timestamps = self._process_with_vad(audio_int16, sr, vad)

        # Filter segments by duration and extract audio data
        segments_data = []

        for idx, (start_sec, end_sec) in enumerate(timestamps, start=1):
            duration = end_sec - start_sec

            # Filter by duration
            if duration < min_duration or duration > max_duration:
                continue

            # Extract segment audio
            start_sample = int(start_sec * sr)
            end_sample = int(end_sec * sr)
            segment_audio = audio[start_sample:end_sample]

            segments_data.append({
                "audio": segment_audio,
                "sample_rate": sr,
                "index": idx
            })

        return segments_data

    def _save_segments(self, segments_data, model, sample_num):
        """Save segment data to files with assigned sample_num"""
        output_dir = self.output_segments_dir / model / str(sample_num)
        output_dir.mkdir(parents=True, exist_ok=True)

        segment_paths = []

        # IMPORTANT: Use enumerate to renumber from 1, not original VAD index
        for seg_idx, seg_data in enumerate(segments_data, start=1):
            output_filename = f"{seg_idx}.wav"
            output_path = output_dir / output_filename
            sf.write(output_path, seg_data['audio'], seg_data['sample_rate'])
            segment_paths.append(output_path)

        return segment_paths

    def _process_with_vad(self, audio_int16, sample_rate, vad):
        """Process audio with TEN VAD"""
        hop_size = vad.hop_size
        total_samples = len(audio_int16)
        num_chunks = total_samples // hop_size

        voice_active = False
        current_start = None
        raw_segments = []

        for i in range(num_chunks):
            start_idx = i * hop_size
            end_idx = start_idx + hop_size
            chunk = audio_int16[start_idx:end_idx]

            probability, flags = vad.process(chunk)
            is_voice = flags == 1

            if is_voice and not voice_active:
                voice_active = True
                current_start = i * hop_size
            elif not is_voice and voice_active:
                voice_active = False
                current_end = i * hop_size
                start_sec = current_start / sample_rate
                end_sec = current_end / sample_rate
                raw_segments.append((start_sec, end_sec))

        if voice_active and current_start is not None:
            end_sec = total_samples / sample_rate
            start_sec = current_start / sample_rate
            raw_segments.append((start_sec, end_sec))

        # Merge short segments
        merged_segments = self._merge_short_segments(raw_segments)

        return merged_segments

    def _merge_short_segments(self, segments, max_gap=0.5):
        """Merge consecutive segments with small gaps"""
        if not segments:
            return []

        merged = []
        current_start, current_end = segments[0]

        for i in range(1, len(segments)):
            next_start, next_end = segments[i]
            gap = next_start - current_end

            if gap <= max_gap:
                current_end = next_end
            else:
                merged.append((current_start, current_end))
                current_start, current_end = next_start, next_end

        merged.append((current_start, current_end))

        return merged

    def _cleanup_workflow(self):
        """Cleanup after segmentation"""
        print("\n" + "="*60)
        print("CLEANUP")
        print("="*60 + "\n")

        # Confirm deletion
        print("This will delete:")
        print(f"  1. {self.input_videos_dir}/")
        print(f"  2. {self.converted_wav_dir}/")
        print(f"  3. {self.input_metadata_path}")
        print("\nOriginal MP4 files are assumed to be backed up elsewhere.")

        response = input("\nProceed with deletion? (Y/N): ").strip().upper()

        if response != "Y":
            print("Cleanup cancelled")
            return

        # Delete directories
        try:
            if self.input_videos_dir.exists():
                shutil.rmtree(self.input_videos_dir)
                print(f"[OK] Deleted: {self.input_videos_dir}")

            if self.converted_wav_dir.exists():
                shutil.rmtree(self.converted_wav_dir)
                print(f"[OK] Deleted: {self.converted_wav_dir}")

            if self.input_metadata_path.exists():
                self.input_metadata_path.unlink()
                print(f"[OK] Deleted: {self.input_metadata_path}")

            # Recreate empty directories
            self.input_videos_dir.mkdir(parents=True, exist_ok=True)
            self.converted_wav_dir.mkdir(parents=True, exist_ok=True)

            print("\nCleanup complete")

        except Exception as e:
            print(f"ERROR during cleanup: {e}")

    def mode3_validate(self):
        """Mode 3: Validate output_metadata.xlsx against actual files"""
        print("\n" + "="*60)
        print("MODE 3: VALIDATE - Metadata Validation")
        print("="*60 + "\n")

        if not self.output_metadata_path.exists():
            print(f"ERROR: {self.output_metadata_path} not found")
            return False

        # Load output metadata
        wb = load_workbook(self.output_metadata_path)
        ws = wb.active

        total_records = ws.max_row - 1
        missing_files = []
        valid_files = 0

        print(f"Validating {total_records} records...\n")

        for row in range(2, ws.max_row + 1):
            # Get data from Excel
            file_name = ws.cell(row=row, column=1).value
            sample_num = ws.cell(row=row, column=3).value
            segment_num = ws.cell(row=row, column=4).value
            path_str = ws.cell(row=row, column=5).value

            # If path is empty, we need to infer model from file structure
            if not path_str or str(path_str).strip() == "":
                # Try to find the file by scanning directories
                found = False
                for model_dir in self.output_segments_dir.iterdir():
                    if model_dir.is_dir():
                        expected_path = model_dir / str(sample_num) / f"{segment_num}.wav"
                        if expected_path.exists():
                            valid_files += 1
                            found = True
                            break

                if not found:
                    missing_files.append(f"<inferred>/{sample_num}/{segment_num}.wav")
            else:
                # Use provided path
                file_path = self.output_segments_dir / path_str

                if file_path.exists():
                    valid_files += 1
                else:
                    missing_files.append(path_str)

        # Print results
        print("="*60)
        print("VALIDATION RESULTS")
        print("="*60)
        print(f"Total records: {total_records}")
        print(f"Valid files: {valid_files}")
        print(f"Missing files: {len(missing_files)}")

        if missing_files:
            print("\nMissing files:")
            for missing in missing_files[:10]:
                print(f"  - {missing}")
            if len(missing_files) > 10:
                print(f"  ... and {len(missing_files) - 10} more")

        print("="*60 + "\n")

        return len(missing_files) == 0

    def mode4_clean(self, clean_all=False, clean_input=False, force=False):
        """Mode 4: Clean output or all generated data"""
        print("\n" + "="*60)
        print("MODE 4: CLEAN - Remove Generated Data")
        print("="*60 + "\n")

        # Determine what to delete
        items_to_delete = []

        if clean_all:
            # Delete everything
            items_to_delete = [
                ("Directory", self.input_videos_dir),
                ("Directory", self.converted_wav_dir),
                ("Directory", self.output_segments_dir),
                ("File", self.input_metadata_path),
                ("File", self.output_metadata_path)
            ]
            print("Target: ALL generated data")
        elif clean_input:
            # Delete input data only
            items_to_delete = [
                ("Directory", self.converted_wav_dir),
                ("File", self.input_metadata_path)
            ]
            print("Target: INPUT data only")
        else:
            # Default: Delete output data only
            items_to_delete = [
                ("Directory", self.output_segments_dir),
                ("File", self.output_metadata_path)
            ]
            print("Target: OUTPUT data only")

        # Show what will be deleted
        print("\nThis will delete:")
        exists_count = 0
        for item_type, item_path in items_to_delete:
            if item_path.exists():
                exists_count += 1
                if item_type == "Directory":
                    # Count files in directory
                    file_count = sum(1 for _ in item_path.rglob("*") if _.is_file())
                    print(f"  [DIR] {item_path}/ ({file_count} files)")
                else:
                    print(f"  [FILE] {item_path}")
            else:
                print(f"  [WARNING] {item_path} (not found)")

        if exists_count == 0:
            print("\n[OK] Nothing to delete (all items already missing)")
            return True

        # Confirmation
        if not force:
            print("\n[WARNING] This action cannot be undone!")
            response = input("Continue? (y/N): ").strip().lower()
            if response != 'y':
                print("\n[CANCELLED]")
                return False

        # Delete items
        print("\nDeleting...")
        deleted_count = 0
        error_count = 0

        for item_type, item_path in items_to_delete:
            try:
                if not item_path.exists():
                    continue

                if item_type == "Directory":
                    shutil.rmtree(item_path)
                    print(f"  [OK] Deleted directory: {item_path}")
                    # Recreate empty directory
                    item_path.mkdir(parents=True, exist_ok=True)
                    print(f"  [OK] Recreated empty: {item_path}")
                else:
                    item_path.unlink()
                    print(f"  [OK] Deleted file: {item_path}")

                deleted_count += 1

            except Exception as e:
                print(f"  [ERROR] deleting {item_path}: {e}")
                error_count += 1

        # Summary
        print("\n" + "="*60)
        print("CLEAN COMPLETE")
        print("="*60)
        print(f"Items deleted: {deleted_count}")
        if error_count > 0:
            print(f"Errors: {error_count}")
        print("="*60 + "\n")

        return error_count == 0


def main():
    parser = argparse.ArgumentParser(
        description="Dataset Management Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python manage_dataset.py prepare
  python manage_dataset.py segment
  python manage_dataset.py segment --no-cleanup
  python manage_dataset.py validate
  python manage_dataset.py clean
  python manage_dataset.py clean --all
  python manage_dataset.py clean --force
        """
    )

    parser.add_argument(
        "mode",
        choices=["prepare", "segment", "validate", "clean"],
        help="Operation mode: prepare (mp4→wav), segment (wav→segments), validate (check metadata), clean (remove data)"
    )

    parser.add_argument(
        "--min-duration",
        type=float,
        default=4.0,
        help="Minimum segment duration in seconds (default: 4.0)"
    )

    parser.add_argument(
        "--max-duration",
        type=float,
        default=10.0,
        help="Maximum segment duration in seconds (default: 10.0)"
    )

    parser.add_argument(
        "--no-cleanup",
        action="store_true",
        help="Skip cleanup after segmentation (keep input_videos, converted_wav)"
    )

    parser.add_argument(
        "--all",
        action="store_true",
        help="[clean mode] Delete all generated data (input + output)"
    )

    parser.add_argument(
        "--input",
        action="store_true",
        help="[clean mode] Delete input data only (converted_wav, input_metadata)"
    )

    parser.add_argument(
        "--force",
        action="store_true",
        help="[clean mode] Skip confirmation prompt"
    )

    args = parser.parse_args()

    # Initialize manager
    manager = DatasetManager()

    # Execute mode
    if args.mode == "prepare":
        success = manager.mode1_prepare()
    elif args.mode == "segment":
        success = manager.mode2_segment(
            min_duration=args.min_duration,
            max_duration=args.max_duration,
            no_cleanup=args.no_cleanup
        )
    elif args.mode == "validate":
        success = manager.mode3_validate()
    elif args.mode == "clean":
        success = manager.mode4_clean(
            clean_all=args.all,
            clean_input=args.input,
            force=args.force
        )

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
